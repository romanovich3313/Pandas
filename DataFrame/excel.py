import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

data = pd.read_csv("sales_data1.csv")


pivot_table = data.pivot_table(
    index='category',  
    columns='month',   
    values='total_sales',  
    aggfunc='sum',         
    margins=True,          
    margins_name='Total'   
)


output_file = 'sales_pivot_table.xlsx'
pivot_table.to_excel(output_file, sheet_name='Pivot Table')


wb = openpyxl.load_workbook(output_file)
sheet = wb['Pivot Table']

#форматування
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
rule = CellIsRule(operator="greaterThan", formula=["10000"], fill=fill)
sheet.conditional_formatting.add(f"B2:{sheet.dimensions.split(':')[1]}", rule)


wb.save(output_file)
output_file
